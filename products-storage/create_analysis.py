import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# --- Лист 1: Сравнительная таблица ---
ws1 = wb.active
ws1.title = "Сравнение линеек"

# Заголовки
headers = ["Тип", "№", "Целевое название (стратегия)", "Текущее название", "Статус", "Цена (стратегия)", "Примечания"]
ws1.append(headers)

# Стили заголовков
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

for col in range(1, len(headers)+1):
    cell = ws1.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_align

# Данные: Бесплатная линейка
free_products = [
    ["Бесплатная", 1, "КС подписаны, денег нет: первые 7 проверок", "—", "ОТСУТСТВУЕТ", "0 ₽", "Нужно создать: алгоритм проверки, калькулятор срока, перечень документов, шаблон письма"],
    ["Бесплатная", 2, "Допы не в подарок", "—", "ОТСУТСТВУЕТ", "0 ₽", "Нужно создать: чек-лист, схема фиксации, шаблон уведомления, реестр Excel"],
    ["Бесплатная", 3, "КС без возврата", "—", "ОТСУТСТВУЕТ", "0 ₽", "Нужно создать: проверочный лист, реестр приложений, типовые причины возврата, порядок проверки"],
]

# Данные: Платная линейка
paid_products = [
    ["Платная", 1, "КС подписаны — денег нет", "01-закрытие-работ", "ЕСТЬ (переименовать)", "14 900–19 900 ₽", "Текущий пакет содержит КС-2, КС-3, акты, реестры — соответствует теме. Нужно переименовать и добавить продающее описание."],
    ["Платная", 2, "Допы не в подарок", "02-допработы-без-потерь", "ЕСТЬ (переименовать)", "14 900–19 900 ₽", "Текущий пакет про допработы — соответствует. Нужно переименовать."],
    ["Платная", 3, "КС без возврата", "—", "ОТСУТСТВУЕТ", "12 900–16 900 ₽", "Нет отдельного продукта. Можно выделить из 01 или создать новый."],
    ["Платная", 4, "ИД блокирует оплату", "03-договор-подряда", "ЕСТЬ (частично)", "14 900–19 900 ₽", "Текущий пакет 'Договор подряда' — шире темы ИД. Нужно либо переориентировать, либо создать отдельный продукт под ИД."],
    ["Платная", 5, "Удержания, штрафы и зачёты", "—", "ОТСУТСТВУЕТ", "19 900–24 900 ₽", "Нет аналога в текущей структуре. Создать новый продукт."],
    ["Платная", 6, "Полный комплект 'Защита маржи объекта'", "04-полный-комплект-ПТО", "ЕСТЬ (переименовать)", "49 900–69 900 ₽", "Текущий комплект включает 01-03 + допы. Нужно добавить 04 и 05."],
]

thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

row = 2
for data in free_products + paid_products:
    ws1.append(data)
    for col in range(1, len(headers)+1):
        cell = ws1.cell(row=row, column=col)
        cell.border = thin_border
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if data[4] == "ОТСУТСТВУЕТ":
            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            cell.font = Font(color="9C0006")
        elif "ЕСТЬ" in data[4]:
            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            cell.font = Font(color="006100")
    row += 1

# Ширина колонок
ws1.column_dimensions['A'].width = 14
ws1.column_dimensions['B'].width = 6
ws1.column_dimensions['C'].width = 38
ws1.column_dimensions['D'].width = 28
ws1.column_dimensions['E'].width = 22
ws1.column_dimensions['F'].width = 18
ws1.column_dimensions['G'].width = 60

# --- Лист 2: Карта контента ---
ws2 = wb.create_sheet(title="Карта контента")

content_headers = ["Продукт", "Файл", "Тип", "Назначение", "Рекомендация"]
ws2.append(content_headers)
for col in range(1, len(content_headers)+1):
    cell = ws2.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_align

content_rows = [
    ["01 Закрытие работ", "01-ks-2.docx", "Word", "Форма КС-2", "Оставить"],
    ["01 Закрытие работ", "02-ks-3.docx", "Word", "Форма КС-3", "Оставить"],
    ["01 Закрытие работ", "03-akt-vypolnennyh-rabot.docx", "Word", "Акт выполненных работ", "Оставить"],
    ["01 Закрытие работ", "04-akt-priemki.docx", "Word", "Акт приемки", "Оставить"],
    ["01 Закрытие работ", "05-peredatochnyy-akt.docx", "Word", "Передаточный акт", "Оставить"],
    ["01 Закрытие работ", "06-zhurnal-obemov.xlsx", "Excel", "Журнал объемов", "Оставить"],
    ["01 Закрытие работ", "07-reestr-zamechaniy.xlsx", "Excel", "Реестр замечаний", "Оставить"],
    ["01 Закрытие работ", "08-reestr-peredachi.xlsx", "Excel", "Реестр передачи", "Оставить"],
    ["01 Закрытие работ", "09-checklist-peredachi.pdf", "PDF", "Чек-лист передачи", "Оставить"],
    ["01 Закрытие работ", "10-sroki-hraneniya.pdf", "PDF", "Сроки хранения", "Оставить"],
    ["", "", "", "", ""],
    ["02 Допработы без потерь", "01-prikaz-na-dopobem.docx", "Word", "Приказ на допобъем", "Оставить"],
    ["02 Допработы без потерь", "02-soglasovanie-obema.docx", "Word", "Согласование объема", "Оставить"],
    ["02 Допработы без потерь", "03-uvedomlenie-o-doprabotah.docx", "Word", "Уведомление о допработах", "Оставить"],
    ["02 Допработы без потерь", "04-akt-skrytyh-rabot.docx", "Word", "Акт скрытых работ", "Оставить"],
    ["02 Допработы без потерь", "05-izmenenie-srokov.docx", "Word", "Изменение сроков", "Оставить"],
    ["02 Допработы без потерь", "06-pismo-o-priostanovke.docx", "Word", "Письмо о приостановке", "Оставить"],
    ["02 Допработы без потерь", "07-raschet-stoimosti.xlsx", "Excel", "Расчет стоимости", "Оставить"],
    ["02 Допработы без потерь", "08-zhurnal-doprabot.xlsx", "Excel", "Журнал допработ", "Оставить"],
    ["02 Допработы без потерь", "09-checklist-fotofiksacii.pdf", "PDF", "Чек-лист фотофиксации", "Оставить"],
    ["02 Допработы без потерь", "10-algoritm-doprabot.pdf", "PDF", "Алгоритм допработ", "Оставить"],
    ["", "", "", "", ""],
    ["03 Договор подряда", "01-dogovor-subpodryada.docx", "Word", "Договор субподряда", "Пересмотреть — не соответствует 'ИД блокирует оплату'"],
    ["03 Договор подряда", "04-poryadok-priemki.docx", "Word", "Порядок приемки", "Оставить"],
    ["03 Договор подряда", "06-dopsoglashenie-obem.docx", "Word", "Допсоглашение объем", "Оставить"],
    ["03 Договор подряда", "07-dopsoglashenie-sroki.docx", "Word", "Допсоглашение сроки", "Оставить"],
    ["03 Договор подряда", "10-protokol-raznoglasiy.docx", "Word", "Протокол разногласий", "Оставить"],
    ["03 Договор подряда", "02-perechen-rabot.xlsx", "Excel", "Перечень работ", "Оставить"],
    ["03 Договор подряда", "03-kalendarnyy-plan.xlsx", "Excel", "Календарный план", "Оставить"],
    ["03 Договор подряда", "05-grafik-platezhey.xlsx", "Excel", "График платежей", "Оставить"],
    ["03 Договор подряда", "08-krasnye-flagi.pdf", "PDF", "Красные флаги", "Оставить"],
    ["03 Договор подряда", "09-checklist-dogovora.pdf", "PDF", "Чек-лист договора", "Оставить"],
    ["", "", "", "", ""],
    ["04 Полный комплект ПТО", "31-40 (допы)", "Mixed", "Дополнительные файлы", "Оставить"],
]

row = 2
for data in content_rows:
    ws2.append(data)
    for col in range(1, len(content_headers)+1):
        cell = ws2.cell(row=row, column=col)
        cell.border = thin_border
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    row += 1

ws2.column_dimensions['A'].width = 26
ws2.column_dimensions['B'].width = 32
ws2.column_dimensions['C'].width = 10
ws2.column_dimensions['D'].width = 30
ws2.column_dimensions['E'].width = 50

# --- Лист 3: План действий ---
ws3 = wb.create_sheet(title="План действий")

plan_headers = ["Приоритет", "Действие", "Сложность", "Оценка времени", "Ответственный", "Статус"]
ws3.append(plan_headers)
for col in range(1, len(plan_headers)+1):
    cell = ws3.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_align

plan_rows = [
    ["P0", "Создать 3 бесплатных продукта (lead-magnets)", "Средняя", "2-3 дня", "Контент", "Не начато"],
    ["P0", "Переименовать 01 → 'КС подписаны — денег нет'", "Низкая", "1 час", "Менеджер", "Не начато"],
    ["P0", "Переименовать 02 → 'Допы не в подарок'", "Низкая", "1 час", "Менеджер", "Не начато"],
    ["P1", "Переосмыслить 03: либо 'ИД блокирует оплату', либо новый продукт", "Высокая", "3-5 дней", "Юрист + Контент", "Не начато"],
    ["P1", "Создать новый продукт 'КС без возврата'", "Средняя", "2-3 дня", "Контент", "Не начато"],
    ["P1", "Создать новый продукт 'Удержания, штрафы и зачёты'", "Высокая", "3-5 дней", "Юрист + Контент", "Не начато"],
    ["P1", "Переименовать 04 → 'Защита маржи объекта'", "Низкая", "1 час", "Менеджер", "Не начато"],
    ["P2", "Добавить в каждый бесплатный продукт: ссылку на платный, сравнительную таблицу, превью платного", "Средняя", "1 день", "Маркетинг", "Не начато"],
    ["P2", "Обновить SERVER-MANIFEST.txt с новыми названиями", "Низкая", "30 мин", "Технический", "Не начато"],
    ["P2", "Настроить защищенную выдачу без прямых ссылок", "Низкая", "2 часа", "Технический", "Не начато"],
]

row = 2
for data in plan_rows:
    ws3.append(data)
    for col in range(1, len(plan_headers)+1):
        cell = ws3.cell(row=row, column=col)
        cell.border = thin_border
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    row += 1

ws3.column_dimensions['A'].width = 10
ws3.column_dimensions['B'].width = 55
ws3.column_dimensions['C'].width = 12
ws3.column_dimensions['D'].width = 16
ws3.column_dimensions['E'].width = 20
ws3.column_dimensions['F'].width = 14

# Сохраняем
output_path = r"C:\Users\X6\Documents\Мажа в бетоне\marzhavbetone-products-v1\Анализ_продуктовой_линейки.xlsx"
wb.save(output_path)
print(f"Excel сохранен: {output_path}")
